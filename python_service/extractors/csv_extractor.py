import os
import csv
from datetime import datetime
from openpyxl import load_workbook


def _infer_column_type(values: list) -> str:
    if not values:
        return 'unknown'

    # Check if all values are numeric
    try:
        float_values = [float(v) for v in values if v and v.strip()]
        if len(float_values) > len(values) * 0.8:  # 80% can be converted to float
            # Check if they are integers
            if all(v.is_integer() for v in float_values):
                return 'integer'
            return 'float'
    except (ValueError, AttributeError):
        pass

    # Check if they look like dates
    # Simple check: if most values contain common date separators
    date_indicators = ['-', '/', ':']
    date_like = sum(1 for v in values if any(ind in str(v) for ind in date_indicators))
    if date_like > len(values) * 0.5:
        return 'datetime'

    return 'string'


def _column_metadata(name: str, values: list):
    non_null_count = sum(1 for v in values if v is not None and str(v).strip())
    total = len(values)
    null_count = total - non_null_count
    null_percentage = round((null_count / total) * 100, 4) if total > 0 else 0.0

    sample_values = [str(v) for v in values[:3] if v is not None]

    col_type = _infer_column_type(values)
    cardinality = len(set(str(v).strip().lower() for v in values if v is not None and str(v).strip()))
    unique_percentage = round((cardinality / max(1, non_null_count)) * 100, 2) if non_null_count > 0 else 0.0

    column_info = {
        'name': name,
        'type': col_type,
        'position': None,  # Will be set by caller
        'non_null_count': non_null_count,
        'null_count': null_count,
        'null_percentage': null_percentage,
        'cardinality': cardinality,
        'unique_percentage': unique_percentage,
        'sample_values': sample_values,
    }

    if col_type in ['integer', 'float']:
        try:
            numeric_values = [float(v) for v in values if v is not None and str(v).strip()]
            if numeric_values:
                column_info['numeric_range'] = {
                    'min': min(numeric_values),
                    'max': max(numeric_values),
                    'mean': sum(numeric_values) / len(numeric_values)
                }
        except (ValueError, TypeError):
            pass

    return column_info


def _normalize_rows(raw_rows: list[list]) -> list[list]:
    """Trim trailing empty cells and normalize each row to a list."""
    normalized_rows = []
    for row in raw_rows:
        values = list(row) if row is not None else []
        while values and (values[-1] is None or str(values[-1]).strip() == ''):
            values.pop()
        normalized_rows.append(values)
    return normalized_rows


def _read_delimited_rows(file_path: str, separator: str) -> list[list]:
    rows = []
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        reader = csv.reader(f, delimiter=separator)
        for row in reader:
            rows.append(row)
    return _normalize_rows(rows)


def _read_excel_rows(file_path: str) -> list[list]:
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return _normalize_rows(rows)


def extract_csv_metadata(file_path: str, file_name: str = None, uploaded_by: str = 'unknown', uploaded_at: str = None):
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f'file does not exist: {file_path}')

    file_name = file_name or os.path.basename(file_path)
    _, ext = os.path.splitext(file_name)
    ext = ext.lower().strip('.')

    if ext not in ['csv', 'tsv', 'xlsx', 'xls']:
        raise ValueError(f'Unsupported extension for metadata extraction: {ext}')

    if ext == 'csv':
        rows = _read_delimited_rows(file_path, ',')
    elif ext == 'tsv':
        rows = _read_delimited_rows(file_path, '\t')
    else:
        rows = _read_excel_rows(file_path)

    if not rows:
        raise ValueError('Source file is empty')

    # First row is headers
    headers = rows[0]
    data_rows = rows[1:]

    row_count = len(data_rows)
    column_count = len(headers)

    # Transpose data to get columns
    columns_data = []
    for i in range(column_count):
        column_values = []
        for row in data_rows:
            if i < len(row):
                column_values.append(row[i])
            else:
                column_values.append(None)
        columns_data.append(column_values)

    # Extract metadata for each column
    columns_info = []
    for i, (header, values) in enumerate(zip(headers, columns_data)):
        col_metadata = _column_metadata(header, values)
        col_metadata['position'] = i
        columns_info.append(col_metadata)

    # Calculate quality metrics
    type_counts = {}
    for col in columns_info:
        col_type = col.get('type', 'unknown')
        type_counts[col_type] = type_counts.get(col_type, 0) + 1

    completeness = round((sum(c['non_null_count'] for c in columns_info) / (max(1, row_count) * column_count)) * 100, 4) if column_count > 0 else 0.0

    metadata = {
        'file_info': {
            'file_name': file_name,
            'file_type': ext,
            'file_size_bytes': int(os.path.getsize(file_path)),
            'uploaded_at': uploaded_at,
            'uploaded_by': uploaded_by
        },
        'dimensions': {
            'row_count': row_count,
            'column_count': column_count
        },
        'columns': columns_info,
        'quality_metrics': {
            'completeness': completeness,
            'data_type_distribution': type_counts,
            'extraction_version': '1.0'
        },
        'metadata_extracted_at': datetime.utcnow().isoformat() + 'Z'
    }

    return metadata
